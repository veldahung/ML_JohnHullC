from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os
import os.path as osp
from IPython.display import display_html
from sklearn.metrics import confusion_matrix
from urllib.request import urlretrieve
import matplotlib.pyplot as plt
import seaborn as sns


def read_excel_in_range(ch_path, fname, sheet_name, start_cell, end_cell):
    file_path = osp.join(ch_path, fname)
    wb = load_workbook(file_path, read_only=True)
    ws = wb[sheet_name]

    data_rows = []
    for row in ws[start_cell:end_cell]:
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)
        
    return pd.DataFrame(data_rows[1:], columns=data_rows[0])


def draw_corr_heatmap(cov_matrix, cols):
    plt.figure(figsize=(6,6), facecolor='w')
    sns.set(font_scale=1.2)
    hm = sns.heatmap(cov_matrix,
                     cmap='coolwarm',
                     annot=True,
                     square=True,
                     fmt='.2f',
                     yticklabels=cols,
                     xticklabels=cols)
    plt.tight_layout()
    plt.show()


def display_side_by_side(dfs: list, names=None, descriptions=None):
    if names is None:
        names = []
    html_str = ''
    if names is not None:
        html_str += ('<tr>' + 
                     ''.join(f'<td style="text-align:center">{name}</td>' for name in names) + 
                     '</tr>')

    html_str += ('<tr>' + 
                 ''.join(f'<td style="vertical-align:top"> {df.to_html()}</td>' 
                         for df in dfs) + 
                 '</tr>')

    if descriptions is not None:
        html_str += ('<tr>' +
                     ''.join(f'<td>{description}</td>' for description in descriptions) +
                     '</tr>')
    html_str = f'<table>{html_str}</table>'
    html_str = html_str.replace('table', 'table style="display:inline"')
    display_html((html_str, ), raw=True)


def classifier_measurement_under_threshs(y_true, y_score, threshs: list):
    ratios = pd.DataFrame(columns=threshs, index=['Accuracy', 'TPR', 'TNR', 'FPR', 'Precision', 'F-Score',
                                                  'F_0.5-Score', 'F_2-Score',
                                                  'tp', 'fn', 'fp', 'tn'])
    for threshold in threshs:
        is_greater = (pd.DataFrame(y_score)[1] > threshold).astype(int)
        tp, fn, fp, tn = np.flip(confusion_matrix(y_true, is_greater, normalize='all')).ravel() * 100
        accuracy = (tp+tn)/(tp+fn+fp+tn) * 100
        TPR = tp/(tp+fn) * 100
        TNR = tn/(fp+tn) * 100
        FPR = fp/(fp+tn) * 100
        precision = tp/(tp+fp) * 100
        fscore = 2*precision*TPR/(precision+TPR)
        f05score = 1.25*precision*TPR/(0.25*precision+TPR)
        f2score = 5*precision*TPR/(4*precision+TPR)
        ratios.loc[:, threshold] = [accuracy, TPR, TNR, FPR, precision, fscore, f05score, f2score, tp, fn, fp, tn]

    return ratios.round(2)


class DataLoader:
    data = {
        "iowa_reduced_scaled": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Iowa%20House%20Price"
                   "%20Regression%20(Ch3)/Houseprice_data_scaled.csv",
            "description": "iowa house price data with features selected & scaled."
        },
        "iowa_original": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Iowa%20House%20Price"
                   "%20Regression%20(Ch3)/Original_Data.xlsx",
            "description": "iowa house price data with all features & not scaled.",
            "excel_kwargs": {
                "sheet_name": "Sheet1",
                "start_cell": "A4",
                "end_cell": "CC2912"
            }
        },
        "age_salary_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Salary%20vs%20Age"
                   "%20Example%20(Chs%201,%203,%206)/Salary%20vs.%20Age%20Example.xlsx",
            "description": "age salary data with train/val/test scattered on different sheet.",
            "excel_kwargs": {
                "train": {
                    "sheet_name": "Analysis of Training Set (Ch1)",
                    "start_cell": "B3",
                    "end_cell": "C13"
                },
                "val": {
                    "sheet_name": "Fit to Validation Set (Ch1)",
                    "start_cell": "B3",
                    "end_cell": "C13"
                },
                "test": {
                    "sheet_name": "Test Set Errors (Ch1)",
                    "start_cell": "B2",
                    "end_cell": "C12"
                }
            }
        },
        "lending_club_reduced_train_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Lending%20Club%20Logistic"
                   "%20Regression%20(Ch%203)/lendingclub_traindata.xlsx",
            "description": "lending club train data with features selected.",
            "excel_kwargs": {
                "sheet_name": "Sheet1",
                "start_cell": "A1",
                "end_cell": "E8696"
            }
        },
        "lending_club_reduced_test_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Lending%20Club%20Logistic"
                   "%20Regression%20(Ch%203)/lendingclub_testdata.xlsx",
            "description": "lending club test data with features selected.",
            "excel_kwargs": {
                "sheet_name": "Sheet1",
                "start_cell": "A1",
                "end_cell": "E8696"
            }
        },
        "lending_club_full_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Lending%20Club%20Logistic"
                   "%20Regression%20(Ch%203)/lending_clubFull_Data_Set.xlsx",
            "description": "lending club all features full data",
            "excel_kwargs": {
                "sheet_name": "Sheet1",
                "start_cell": "A1",
                "end_cell": "EE25001"
            }
        },
        "country_risk_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Country%20risk%20Example"
                   "%20(Ch2)/countryriskdata.csv",
            "description": "country risk data",
        },
        "new_2019_country_risk_url": {
            "url": "http://www-2.rotman.utoronto.ca/~hull/Second%20edition%20Online%20Files/Country%20risk%20Example"
                   "%20(Ch2)/Country%20Risk%202019%20Data.xlsx",
            "description": "new 2019 country risk data",
            "excel_kwargs": {
                "sheet_name": "Sheet1",
                "start_cell": "A2",
                "end_cell": "E124"
            }
        },
    }

    @classmethod
    def list_available_data(cls):
        for name in cls.data:
            print("\t", name)

    @classmethod
    def show_description(cls, name):
        print(cls.data[name]['description'])

    @classmethod
    def load_data(cls, name: str, sheet_name=''):
        if name not in cls.data:
            print("{} is not registered!".format(name))
            return None

        if not os.path.exists('data'):
            os.makedirs('data')

        type_of_the_file = cls.data[name]['url'].split('.')[-1]
        file_name = '{}.{}'.format(name, type_of_the_file)
        data_path = osp.join('data', file_name)
        if not os.path.exists(data_path):
            _ = urlretrieve(cls.data[name]['url'], data_path)
        if type_of_the_file == 'csv':
            return pd.read_csv(data_path)
        elif type_of_the_file == 'xlsx' and sheet_name:
            return read_excel_in_range('data', file_name, **cls.data[name]['excel_kwargs'][sheet_name])
        elif type_of_the_file == 'xlsx':
            return read_excel_in_range('data', file_name, **cls.data[name]['excel_kwargs'])
        else:
            return None
